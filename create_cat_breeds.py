from openpyxl import Workbook

# Create a new workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Cat Breeds"

# Add headers
ws['A1'] = 'Breed Name'
ws['B1'] = 'Origin'
ws['C1'] = 'Temperament'
ws['D1'] = 'Size'
ws['E1'] = 'Life Expectancy'

# Add cat breeds data
cat_breeds = [
    ['Persian', 'Iran', 'Gentle and quiet', 'Medium to large', '12-17 years'],
    ['Siamese', 'Thailand', 'Active and social', 'Medium', '12-20 years'],
    ['Maine Coon', 'United States', 'Friendly and playful', 'Large', '12-15 years'],
    ['British Shorthair', 'United Kingdom', 'Calm and affectionate', 'Medium to large', '12-17 years'],
    ['Bengal', 'United States', 'Energetic and playful', 'Medium to large', '12-16 years']
]

# Add data to worksheet
for row, breed in enumerate(cat_breeds, start=2):
    for col, value in enumerate(breed):
        ws.cell(row=row, column=col+1, value=value)

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    ws.column_dimensions[column].width = max_length + 2

# Save the workbook
wb.save('cat_breeds.xlsx')