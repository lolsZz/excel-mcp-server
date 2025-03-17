from openpyxl import load_workbook

def read_dogs_excel():
    # Load the workbook and select the active sheet
    wb = load_workbook('dog_breeds.xlsx')
    ws = wb.active

    # Print sheet information
    print(f"\nReading sheet: {ws.title}\n")

    # Get max row and column
    max_row = ws.max_row
    max_col = ws.max_column

    # Print headers
    headers = []
    for col in range(1, max_col + 1):
        header = ws.cell(row=1, column=col).value
        headers.append(header)
    
    # Calculate column widths
    col_widths = [len(str(header)) for header in headers]
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell_value = str(ws.cell(row=row, column=col).value)
            col_widths[col-1] = max(col_widths[col-1], len(cell_value))

    # Print formatted table
    # Print header separator
    print("+" + "+".join("-" * (width + 2) for width in col_widths) + "+")
    
    # Print headers
    header_row = "|"
    for header, width in zip(headers, col_widths):
        header_row += f" {header:{width}} |"
    print(header_row)
    
    # Print separator
    print("+" + "+".join("-" * (width + 2) for width in col_widths) + "+")
    
    # Print data
    for row in range(2, max_row + 1):
        data_row = "|"
        for col, width in enumerate(col_widths, 1):
            value = str(ws.cell(row=row, column=col).value)
            data_row += f" {value:{width}} |"
        print(data_row)
    
    # Print bottom separator
    print("+" + "+".join("-" * (width + 2) for width in col_widths) + "+")

if __name__ == "__main__":
    read_dogs_excel()