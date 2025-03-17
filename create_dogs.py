from openpyxl import Workbook

# Dog breeds data
dog_data = {
    "headers": ["Breed Name", "Origin", "Temperament", "Size", "Life Expectancy"],
    "data": [
        ["Labrador Retriever", "Canada", "Friendly and outgoing", "Large", "10-12 years"],
        ["German Shepherd", "Germany", "Confident and courageous", "Large", "9-13 years"],
        ["Golden Retriever", "Scotland", "Intelligent and friendly", "Large", "10-12 years"],
        ["French Bulldog", "France", "Playful and adaptable", "Small", "10-12 years"],
        ["Beagle", "England", "Merry and friendly", "Medium", "12-15 years"]
    ]
}

def create_dogs_excel():
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Dog Breeds"

    # Write headers
    for col, header in enumerate(dog_data["headers"], 1):
        ws.cell(row=1, column=col, value=header)

    # Write data
    for row, breed_info in enumerate(dog_data["data"], 2):
        for col, value in enumerate(breed_info, 1):
            ws.cell(row=row, column=col, value=value)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column].width = max_length + 2

    # Save the workbook
    wb.save('dog_breeds.xlsx')
    print("Dog breeds Excel file created successfully!")

if __name__ == "__main__":
    create_dogs_excel()